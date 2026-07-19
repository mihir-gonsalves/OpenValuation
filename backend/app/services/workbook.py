# backend/app/services/workbook.py
"""
Excel workbook builder - Phase 4.

Turns a computed `FinancialsResponse` into a downloadable `.xlsx` whose every
result is a live formula, never a hardcoded number.  Editing any reported input
recomputes that period's market cap, enterprise value, and seven multiples on the
spot - the workbook is the audit trail *and* a scratchpad.

Figures are quoted "USD (values in thousands, except per share)".  Multiples are 
unchanged - each is a ratio of two thousands-scaled figures. See `_SCALE`.

Layout (chosen so the reader is never staring at a wall of numbers):

  * `Summary`           - Company metadata, valuation metrics, a seven multiple 
                          matrix across every period, and formula-based links 
                          into each period sheet.
  * `{period_end}`      - One sheet per TTM period. Three stacked blocks:
        Inputs            - Every reported XBRL value with its tag, fallback status,
                            and unit.
        Calculations      - Market Cap and the Enterprise Value buildup as formulas
                            referencing the Inputs block.
        Multiples         - The seven valuation multiples as formulas, each reproducing
                            the backend guards (missing data, near-zero denominator,
                            negative book value / FCF) so an edit yields the same N/A
                            the backend would.
        Notes             - The period's data-quality warnings.

The financial semantics reproduced by the formulas are authoritative in
`README.md` / `services/multiples.py`, this module only mirrors them for Excel.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.financials import FinancialsResponse, TTMPeriod

# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------

_PRICE = "$#,##0.00;($#,##0.00)"
_COUNT = "#,##0;(#,##0)"
_MULT = "0.00\"x\";(0.00\"x\")"
_DATE = "yyyy-mm-dd"

_SCALE = Decimal(1000)

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

_INK = "284876"        # --primary
_MUTED = "67625D"      # --muted-foreground
_AMBER = "E5AC4C"      # --warning
_BLACK = "000000"      # formula
_BLUE = "0000FF"       # hardcoded input
_GREEN = "008000"      # link to another sheet

_SECTION_FONT = Font(bold=True, size=12, color=_INK)
_NOTE_FONT = Font(italic=True, size=9, color=_MUTED)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor=_INK)
_LABEL_FONT = Font(bold=True)
_FALLBACK_FONT = Font(color=_AMBER, bold=True)
_INPUT_FONT = Font(color=_BLUE)                       # blue: editable input
_LINK_FONT = Font(color=_GREEN)                       # green: cross-sheet link

_RIGHT = Alignment(horizontal="right")
_LEFT = Alignment(horizontal="left")
_WRAP = Alignment(vertical="top", wrap_text=True)

_DIVIDER_BORDER = Border(bottom=Side(style="thin", color=_INK))

# ---------------------------------------------------------------------------
# Input groups: (group header, [(display label, ExtractedFinancials attr,
#                                audit concept, format), ...])
#
# The audit concept keys mirror the names xbrl.py assigns to each AuditEntry.
# Order matches the audit list in xbrl.py: market header (price/shares/EPS),
# income statement -> cash flow -> balance sheet.
#
# The debt/cash items mirror the EV buildup order (see compute_enterprise_value
# and the frontend EVBreakdown) so the schedule reconciles line-by-line into the EV bridge.
# ---------------------------------------------------------------------------

_INPUT_GROUPS: list[tuple[str, list[tuple[str, str, str | None, str]]]] = [
    ("Valuation", [
        ("Price", "price", None, _PRICE),
        ("Shares Outstanding", "shares_outstanding", "Shares Outstanding", _COUNT),
        ("EPS", "eps_diluted", "EPS", _PRICE),
    ]),
    ("Income Statement", [
        ("Revenue", "revenue", "Revenue", _COUNT),
        ("Operating Income (EBIT)", "operating_income", "Operating Income", _COUNT),
        ("Depreciation & Amortization", "depreciation_and_amortization", "Depreciation & Amortization", _COUNT),
        ("Net Income", "net_income", "Net Income", _COUNT),
    ]),
    ("Cash Flow", [
        ("Operating Cash Flow", "operating_cash_flow", "Operating Cash Flow", _COUNT),
        ("CapEx", "capex", "CapEx", _COUNT),
    ]),
    ("Balance-Sheet", [
        ("Total Assets", "total_assets", "Total Assets", _COUNT),
        ("Stockholders' Equity", "stockholders_equity", "Stockholders Equity", _COUNT),
        ("Long-Term Debt", "long_term_debt", "Long-Term Debt", _COUNT),
        ("Short-Term Borrowings", "short_term_borrowings", "Short-Term Borrowings", _COUNT),
        ("Current Portion LT Debt", "current_portion_lt_debt", "Current Portion LT Debt", _COUNT),
        ("Finance Lease (Current)", "finance_lease_current", "Finance Lease (Current)", _COUNT),
        ("Finance Lease (Non-Current)", "finance_lease_noncurrent", "Finance Lease (Non-Current)", _COUNT),
        ("Minority Interest", "minority_interest", "Minority Interest", _COUNT),
        ("Preferred Stock", "preferred_stock", "Preferred Stock", _COUNT),
        ("Cash & Equivalents", "cash", "Cash", _COUNT),
    ]),
]

# Summary matrix rows: (label, ref key into the per-period refs dict, format).
# None is a spacer row.
_SUMMARY_ROWS: list[tuple[str, str, str] | None] = [
    ("Price", "price", _PRICE),
    ("Shares Outstanding", "shares_outstanding", _COUNT),
    ("Market Cap", "market_cap", _COUNT),
    ("Enterprise Value", "enterprise_value", _COUNT),
    None,
    ("EV/Revenue", "ev_revenue", _MULT),
    ("EV/EBITDA", "ev_ebitda", _MULT),
    ("EV/EBIT", "ev_ebit", _MULT),
    ("P/E", "pe", _MULT),
    ("P/FCF", "pfcf", _MULT),
    ("P/S", "ps", _MULT),
    ("P/B", "pb", _MULT),
]


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def build_workbook(response: FinancialsResponse) -> bytes:
    """
    Build and return a `.xlsx` workbook as raw bytes, ready to stream as a
    binary HTTP response.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet, name it manually

    summary = wb.create_sheet("Summary")

    # Build the period sheets first, collecting the cell addresses the Summary
    # matrix will point at. Sheet order becomes Summary, then periods.
    period_refs: list[tuple[TTMPeriod, str, dict[str, str]]] = []
    used_names: set[str] = set()
    for period in response.periods:
        name = _period_sheet_name(period.period_end, used_names)
        ws = wb.create_sheet(name)
        refs = _build_period_sheet(ws, period)
        period_refs.append((period, name, refs))

    _build_summary(summary, response, period_refs)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------


def _build_summary(
    ws: Worksheet,
    response: FinancialsResponse,
    period_refs: list[tuple[TTMPeriod, str, dict[str, str]]],
) -> None:
    ws.sheet_view.showGridLines = False
    company = response.company

    ws["A1"] = "OpenValuation - Summary"
    ws["A1"].font = _SECTION_FONT

    meta = [
        ("Company Name", company.name),
        ("Ticker", company.ticker),
        ("Exchange", company.exchange),
        ("CIK", company.cik_10),
        ("SIC", _join(company.sic, company.sic_description)),
        ("Sector", _sector_flags(company)),
        ("Currency / Scale", "USD, in thousands (except per share)"),
        ("Data As Of", response.data_as_of),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=_datetime_or(value))
        cell.alignment = _LEFT
        row += 1

    row += 1
    note = ws.cell(
        row=row, column=1,
        value=("Every result below is a live formula pointing at a sheet. "
               "Click a quarter header to open its inputs, formulas, and warnings. "
               "Edit any reported value there and the multiples recompute. "
               "N/A = data unavailable or a guard fired (e.g. near-zero denominator, negative book value / FCF)."),
    )
    note.font = _NOTE_FONT
    note.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 26
    row += 2

    # Standard color-coding key.
    ws.cell(row=row, column=1, value="Number color key:").font = _LABEL_FONT
    ws.cell(row=row, column=2, value="blue = input").font = _INPUT_FONT
    ws.cell(row=row + 1, column=2, value="black = formula")
    ws.cell(row=row + 2, column=2, value="green = link").font = _LINK_FONT
    row += 4

    if not period_refs:
        ws.cell(row=row, column=1,
                value="No TTM periods were available for this company.").font = _NOTE_FONT
        _set_widths(ws, {"A": 18, "B": 36})
        return

    # --- Matrix header: Metric | one linked column per period -----------------
    header_row = row
    hdr = ws.cell(row=header_row, column=1, value="Metric")
    _apply_header(hdr)
    for col_offset, (period, name, _refs) in enumerate(period_refs, start=2):
        cell = ws.cell(row=header_row, column=col_offset)
        cell.value = _internal_link(name, "A2", name)
        cell.font = Font(bold=True, color="FFFFFF", underline="single")
        cell.fill = _HEADER_FILL
        cell.alignment = _RIGHT
    row += 1

    # --- Matrix body ----------------------------------------------------------
    for spec in _SUMMARY_ROWS:
        if spec is None:
            row += 1
            continue
        label, key, fmt = spec
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        for col_offset, (_period, name, refs) in enumerate(period_refs, start=2):
            cell = ws.cell(row=row, column=col_offset)
            cell.value = f"='{name}'!{refs[key]}"
            cell.number_format = fmt
            cell.alignment = _RIGHT
            cell.font = _LINK_FONT
        row += 1

    # Freeze the metric column
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate

    widths = {get_column_letter(col): 18 for col in range(1, len(period_refs) + 2)}
    _set_widths(ws, widths)


# ---------------------------------------------------------------------------
# Per-period sheet
# ---------------------------------------------------------------------------


def _build_period_sheet(ws: Worksheet, period: TTMPeriod) -> dict[str, str]:
    """
    Write one period sheet and return the cell addresses the Summary needs:
    the price / shares input cells, the market-cap and EV formula cells, and the
    seven multiple formula cells (keyed by MultipleSet field name).
    """
    ws.sheet_view.showGridLines = False
    ef = period.extracted
    audit = {a.concept: a for a in ef.audit}

    ws["A1"] = _internal_link("Summary", "A1", "← Back to Summary")
    ws["A1"].font = _LINK_FONT
    ws["A2"] = f'="{_quarter_label(period.period_end)} - " & Summary!B3'
    ws["A2"].font = _SECTION_FONT

    _kv(ws, 4, "Quarter End", period.period_end, _DATE)
    _kv(ws, 5, "Report Filed", period.filing_date, _DATE)
    _kv(ws, 6, "Currency / Scale", "USD, in thousands (except per share)", _DATE)

    # --- Inputs block ---------------------------------------------------------
    _section(ws, 8, "Inputs")
    sub = ws.cell(
        row=9, column=1,
        value=("Edit any figure in the Value column, the calculations below "
               "recompute automatically. Balance-sheet items left blank are "
               "treated as zero."),
    )
    sub.font = _NOTE_FONT
    sub.alignment = _WRAP
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=4)

    _table_header(ws, 10, ["Concept", "Reported XBRL Tag", "Fallback", "Unit", "Value"])

    refs: dict[str, str] = {}
    r = 11
    for group_idx, (group_title, specs) in enumerate(_INPUT_GROUPS):
        if group_idx > 0:
            r += 1  # blank spacer row between groups
        _table_divider(ws, r, group_title)
        r += 1
        for label, attr, concept, fmt in specs:
            entry = audit.get(concept) if concept else None
            tag = entry.xbrl_tag if entry else ("yfinance (next trading day adj. close)" if attr == "price" else None)
            unit = entry.unit if entry else ("USD/share" if attr == "price" else None)
            is_fallback = bool(entry and entry.is_fallback)

            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=tag or "-")
            fb = ws.cell(row=r, column=3, value="fallback" if is_fallback else "")
            if is_fallback:
                fb.font = _FALLBACK_FONT
            ws.cell(row=r, column=4, value=unit or "-")

            raw = getattr(ef, attr)
            if raw is not None and attr not in {"price", "eps_diluted"}:
                raw = raw / _SCALE
            value_cell = ws.cell(row=r, column=5, value=_num(raw))
            value_cell.number_format = fmt
            value_cell.alignment = _RIGHT
            value_cell.font = _INPUT_FONT
            refs[attr] = value_cell.coordinate
            r += 1

    # --- Calculations block ---------------------------------------------------
    r += 1
    _section(ws, r, "Calculations")
    r += 1
    ws.cell(
        row=r, column=1,
        value="Formulas reference the input values above. No result is hardcoded.",
    ).font = _NOTE_FONT
    r += 1
    _table_header(ws, r, ["Valuation", "Definition", "Value"])
    r += 1

    price, shares = refs["price"], refs["shares_outstanding"]
    mc = ws.cell(row=r, column=3).coordinate
    _calc_row(
        ws,
        r,
        "Market Cap",
        "Price × Shares Outstanding",
        f'=IF(OR({price}="",{shares}=""),"-",{price}*{shares})',
        _COUNT,
    )
    refs["market_cap"] = mc
    r += 1

    ltd, stb, cpltd = refs["long_term_debt"], refs["short_term_borrowings"], refs["current_portion_lt_debt"]
    flc, flnc = refs["finance_lease_current"], refs["finance_lease_noncurrent"]
    mi, pfd, cash = refs["minority_interest"], refs["preferred_stock"], refs["cash"]
    ev = ws.cell(row=r, column=3).coordinate
    _calc_row(
        ws,
        r,
        "Enterprise Value",
        "Mkt Cap + Debt + Leases + Minority + Preferred − Cash",
        f'=IF({mc}="-","-",{mc}+{ltd}+{stb}+{cpltd}+{flc}+{flnc}+{mi}+{pfd}-{cash})',
        _COUNT,
    )
    refs["enterprise_value"] = ev
    r += 1

    # --- Multiples block ------------------------------------------------------
    r += 1
    _table_header(ws, r, ["Multiple", "Definition", "Value"])
    r += 1

    eps, rev, oi, da = refs["eps_diluted"], refs["revenue"], refs["operating_income"], refs["depreciation_and_amortization"]
    ocf, capex, eq = refs["operating_cash_flow"], refs["capex"], refs["stockholders_equity"]

    ms = period.multiples
    multiple_rows = [
        (
            "ev_revenue",
            ms.ev_revenue.label,
            "EV ÷ Revenue",
            f'=IF(OR({ev}="-",{rev}=""),"-",IF(ABS({rev})<0.01,"-",{ev}/{rev}))',
        ),
        (
            "ev_ebitda",
            ms.ev_ebitda.label,
            "EV ÷ (Operating Income + D&A)",
            f'=IF(OR({ev}="-",{oi}="",{da}=""),"-",IF(ABS({oi}+{da})<0.01,"-",{ev}/({oi}+{da})))',
        ),
        (
            "ev_ebit",
            ms.ev_ebit.label,
            "EV ÷ Operating Income",
            f'=IF(OR({ev}="-",{oi}=""),"-",IF(ABS({oi})<0.01,"-",{ev}/{oi}))',
        ),
        (
            "pe",
            ms.pe.label,
            "Price ÷ EPS",
            f'=IF(OR({price}="",{eps}=""),"-",IF(ABS({eps})<0.01,"-",{price}/{eps}))',
        ),
        (
            "pfcf",
            ms.pfcf.label,
            "Market Cap ÷ (Operating Cash Flow − CapEx), N/A if negative",
            f'=IF(OR({mc}="-",{ocf}="",{capex}=""),"-",'
            f'IF(({ocf}-{capex})<0,"-",IF(ABS({ocf}-{capex})<0.01,"-",{mc}/({ocf}-{capex}))))',
        ),
        (
            "ps",
            ms.ps.label,
            "Market Cap ÷ Revenue",
            f'=IF(OR({mc}="-",{rev}=""),"-",IF(ABS({rev})<0.01,"-",{mc}/{rev}))',
        ),
        (
            "pb",
            ms.pb.label,
            "Market Cap ÷ Stockholders' Equity (N/A if negative)",
            f'=IF(OR({mc}="-",{eq}=""),"-",IF({eq}<0,"-",IF(ABS({eq})<0.01,"-",{mc}/{eq})))',
        ),
    ]
    for key, label, formula, definition in multiple_rows:
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=r, column=2, value=formula).font = _NOTE_FONT
        cell = ws.cell(row=r, column=3, value=definition)
        cell.number_format = _MULT
        cell.alignment = _RIGHT
        refs[key] = cell.coordinate
        r += 1

    # --- Notes / warnings block ----------------------------------------------
    r += 1
    _section(ws, r, "Notes/Warnings")
    r += 1
    if period.warnings:
        _table_header(ws, r, ["Code", "Message"])
        r += 1
        for w in period.warnings:
            ws.cell(row=r, column=1, value=w.code).font = _LABEL_FONT
            msg = ws.cell(row=r, column=2, value=w.message)
            msg.alignment = _WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1
    else:
        ws.cell(row=r, column=1,
                value="No data-quality warnings for this period.").font = _NOTE_FONT

    _set_widths(ws, {"A": 30, "B": 45, "C": 15, "D": 15, "E": 15})
    return refs


# ---------------------------------------------------------------------------
# Small writers / formatters
# ---------------------------------------------------------------------------


def _num(x: Decimal | None) -> float | None:
    """Decimal -> float for the cell. Excel stores doubles, this matches it."""
    return float(x) if x is not None else None


def _datetime_or(x):
    """
    openpyxl accepts date/datetime directly, but Excel rejects tz-aware
    datetimes. API's timestamps are UTC, drop the tzinfo (label stays "UTC").
    """
    if x is None:
        return "-"
    if isinstance(x, datetime):
        # Store as text (not a date value) so Excel lets it overflow into
        # adjacent empty cells instead of showing "####" when the column is narrow.
        return x.strftime("%Y-%m-%d %H:%M UTC")
    return x


def _kv(ws: Worksheet, row: int, label: str, value, fmt: str) -> None:
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    cell = ws.cell(row=row, column=2, value=value if value is not None else "-")
    cell.alignment = _LEFT
    if isinstance(value, (date, datetime)):
        cell.number_format = fmt


def _section(ws: Worksheet, row: int, title: str) -> None:
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT


def _calc_row(ws: Worksheet, row: int, label: str, definition: str, formula: str, fmt: str) -> None:
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    ws.cell(row=row, column=2, value=definition).font = _NOTE_FONT
    cell = ws.cell(row=row, column=3, value=formula)
    cell.number_format = fmt
    cell.alignment = _RIGHT


def _table_header(ws: Worksheet, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        _apply_header(cell)
        if label == "Value":
            cell.alignment = _RIGHT


def _table_divider(ws: Worksheet, row: int, label: str) -> None:
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = _DIVIDER_BORDER


def _apply_header(cell) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL


def _internal_link(sheet: str, cell: str, label: str) -> str:
    """A HYPERLINK formula to another sheet in this workbook."""
    label = label.replace('"', "'")
    return f'=HYPERLINK("#\'{sheet}\'!{cell}","{label}")'


def _quarter_label(period_end: date) -> str:
    """date -> 'Q1 2026', the calendar quarter of the period end."""
    quarter = (period_end.month - 1) // 3 + 1
    return f"Q{quarter} {period_end.year}"


def _period_sheet_name(period_end: date, used: set[str]) -> str:
    """'Q1 2026', truncated to Excel's 31-char limit and de-duplicated."""
    base = f"{_quarter_label(period_end)}"[:31]
    name, i = base, 2
    while name in used:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _join(*parts) -> str:
    return " - ".join(str(p) for p in parts if p) or "-"


def _sector_flags(company) -> str:
    flags = []
    if company.is_financial:
        flags.append("Financial")
    if company.is_capital_intensive:
        flags.append("Capital Intensive")
    return ", ".join(flags) if flags else " - "


def _set_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
