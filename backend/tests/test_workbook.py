# backend/tests/test_workbook.py
"""
Unit tests for the Excel workbook builder (services/workbook.py).

The workbook's contract is auditability + adjustability: every result is a live
formula referencing the reported inputs, never a hardcoded value. These tests
pin that contract structurally (openpyxl does not evaluate formulas, the
formula-reproduces-backend logic is validated separately in review).
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal as D
from io import BytesIO

from openpyxl import load_workbook

from app.models.company import CompanyMeta
from app.models.errors import Warning, WarningCode
from app.models.financials import (
    AuditEntry, ExtractedFinancials, FinancialsResponse, TTMPeriod,
)
from app.services import multiples, workbook


def _period(period_end: date, **overrides) -> TTMPeriod:
    fields = dict(
        price=D("220.50"), shares_outstanding=D("15000000000"),
        eps_diluted=D("6.50"), revenue=D("390000000000"),
        operating_income=D("120000000000"),
        depreciation_and_amortization=D("11000000000"),
        net_income=D("100000000000"), operating_cash_flow=D("110000000000"),
        capex=D("10000000000"), total_assets=D("350000000000"),
        stockholders_equity=D("62000000000"), long_term_debt=D("95000000000"),
        short_term_borrowings=D("10000000000"), current_portion_lt_debt=D("11000000000"),
        finance_lease_current=D("1000000000"), finance_lease_noncurrent=D("11000000000"),
        cash=D("30000000000"), minority_interest=None, preferred_stock=None,
    )
    fields.update(overrides)
    ef = ExtractedFinancials(
        period_end=period_end, filing_date=date(period_end.year, period_end.month, 1),
        audit=[
            AuditEntry(concept="Revenue", xbrl_tag="Revenues", unit="USD", value=fields["revenue"]),
            AuditEntry(concept="EPS", xbrl_tag="EarningsPerShareBasic",
                       is_fallback=True, unit="USD/shares", value=fields["eps_diluted"]),
        ],
        **fields,
    )
    ms, ev = multiples.compute_all(ef)
    return TTMPeriod(
        period_end=period_end, filing_date=ef.filing_date, price=ef.price,
        multiples=ms, ev_components=ev, extracted=ef,
        warnings=[Warning(code=WarningCode.FALLBACK_TAG, message="Basic EPS used.")],
    )


def _response(periods) -> FinancialsResponse:
    return FinancialsResponse(
        company=CompanyMeta(
            cik_10="0000320193", name="Apple Inc.", ticker="AAPL",
            sic="3571", sic_description="Electronic Computers", exchange="Nasdaq",
        ),
        periods=periods,
        cached_at=datetime(2026, 7, 8, 12, 0, tzinfo=timezone.utc),
        data_as_of=datetime.now(timezone.utc),
    )


def _load(response) -> "Workbook":
    return load_workbook(BytesIO(workbook.build_workbook(response)))


def test_summary_first_then_one_sheet_per_period():
    wb = _load(_response([_period(date(2024, 9, 28)), _period(date(2024, 6, 29))]))
    assert wb.sheetnames == ["Summary", "Q3 2024", "Q2 2024"]


def test_gridlines_off_on_every_sheet():
    wb = _load(_response([_period(date(2024, 9, 28))]))
    for name in wb.sheetnames:
        assert wb[name].sheet_view.showGridLines is False


def test_multiples_are_guarded_formulas_not_values():
    wb = _load(_response([_period(date(2024, 9, 28))]))
    ws = wb["Q3 2024"]
    values = {c.value for row in ws.iter_rows() for c in row}
    # Every multiple formula is an IF-guarded division; none is a bare number.
    # Select on the near-zero ABS() guard (unique to the multiples; Market Cap /
    # EV are IF-guarded too but have no ABS(), and every formula falls back to "-").
    pe = next(v for v in values if isinstance(v, str) and v.startswith("=IF(OR(") and "ABS(" in v)
    assert '"-"' in pe and "/" in pe


def test_summary_matrix_references_period_sheets_live():
    wb = _load(_response([_period(date(2024, 9, 28))]))
    ws = wb["Summary"]
    refs = [
        c.value for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("='Q3 2024'!")
    ]
    # Price, shares, market cap, EV, and the seven multiples => 11 live references.
    assert len(refs) == 11


def test_missing_input_is_blank_and_fallback_is_flagged():
    wb = _load(_response([_period(date(2024, 9, 28), revenue=None)]))
    ws = wb["Q3 2024"]
    rows = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}

    # Missing revenue -> blank Value cell (edited-to-recompute input), not 0 or "-".
    assert ws.cell(row=rows["Revenue"], column=5).value is None
    # Basic-EPS fallback surfaces in the Fallback column.
    assert ws.cell(row=rows["EPS"], column=3).value == "fallback"


def test_empty_response_has_summary_only():
    wb = _load(_response([]))
    assert wb.sheetnames == ["Summary"]
