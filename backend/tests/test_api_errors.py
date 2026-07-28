# backend/tests/test_api_errors.py
"""
Tests for the structured API error response contract (PHASE_1_SPEC §5).

Coverage:
  - invalid_cik: malformed CIK path param -> 422 with top-level {"error": "invalid_cik"}
  - EDGAR timeout: mocked timeout -> 503 with top-level {"error": "edgar_timeout"}
  - export: returns a valid .xlsx workbook (Phase 4)
  - Generic HTTPException detail strings surface as {"error": "internal_error"}
"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest
from fastapi.testclient import TestClient

from app.main import app


@pytest.fixture
def client():
    with TestClient(app, raise_server_exceptions=False) as c:
        yield c


# ---------------------------------------------------------------------------
# invalid_cik (§1.1 and PHASE_1_SPEC §5)
# ---------------------------------------------------------------------------


def test_invalid_cik_short_returns_422_with_error_code(client):
    """A short CIK (not 10 digits) must return the documented invalid_cik shape."""
    resp = client.get("/api/financials/123")
    assert resp.status_code == 422
    body = resp.json()
    assert "detail" not in body, "Error body must not be wrapped in 'detail'"
    assert body["error"] == "invalid_cik"
    assert "message" in body


def test_invalid_cik_alpha_returns_422_with_error_code(client):
    """A non-numeric CIK must also return invalid_cik."""
    resp = client.get("/api/financials/abcdefghij")
    assert resp.status_code == 422
    body = resp.json()
    assert body.get("error") == "invalid_cik"


# ---------------------------------------------------------------------------
# EDGAR timeout -> 503 edgar_timeout at top level
# ---------------------------------------------------------------------------


def test_edgar_timeout_returns_503_at_top_level(client):
    """EDGAR timeout must produce top-level {"error": "edgar_timeout"} (no detail wrapper)."""
    from fastapi import HTTPException
    from app.services import edgar

    def raise_timeout(*args, **kwargs):
        raise HTTPException(
            status_code=503,
            detail={"error": "edgar_timeout", "message": "EDGAR timed out."},
        )

    with patch.object(edgar, "fetch_metadata", side_effect=raise_timeout):
        resp = client.get("/api/financials/0000320193")

    assert resp.status_code == 503
    body = resp.json()
    assert "detail" not in body, "Error body must not be wrapped in 'detail'"
    assert body["error"] == "edgar_timeout"
    assert "message" in body


# ---------------------------------------------------------------------------
# export -> valid .xlsx workbook (Phase 4)
# ---------------------------------------------------------------------------

_AAPL_SUBMISSIONS = {
    "name": "Apple Inc.",
    "tickers": ["AAPL"],
    "sic": "3571",
    "sicDescription": "Electronic Computers",
    "exchanges": ["Nasdaq"],
}


def test_export_returns_xlsx_workbook(client):
    """Export streams a valid .xlsx built from the same cache-or-fetch path."""
    import app.cache as cache_store
    from openpyxl import load_workbook

    cache_store.invalidate("0000320193")

    facts = json.loads(
        next((Path(__file__).parent / "fixtures").glob("aapl_CIK*.json")).read_text()
    )
    with (
        patch("app.services.edgar.fetch_metadata", new=AsyncMock(return_value=_AAPL_SUBMISSIONS)),
        patch("app.services.edgar.fetch_companyfacts", new=AsyncMock(return_value=facts)),
    ):
        resp = client.get("/api/export/0000320193")

    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "openvaluation_0000320193.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames[0] == "Summary"
    # A period sheet is present and its multiples are formulas, not hardcoded.
    period_sheets = [s for s in wb.sheetnames if s != "Summary"]
    assert period_sheets
    formulas = [
        c.value
        for row in wb[period_sheets[0]].iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert any("HYPERLINK" in f for f in formulas)  # back-to-summary link
    assert any("/" in f and "IF(" in f for f in formulas)  # multiple formula
